import os
import io
from fastapi import FastAPI, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from datetime import datetime
from collections import defaultdict
from dotenv import load_dotenv
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    ExpenseCreate, Expense, ExpenseUpdate, 
    IncomeCreate, IncomeUpdate, AnalyticsResponse,
    BudgetCreate, BudgetUpdate,
    RecurringExpenseCreate, RecurringExpenseUpdate
)
from .database import (
    load_expenses, add_expense, get_expense, update_expense, 
    delete_expense, get_expenses_by_month, get_all_months,
    load_incomes, add_income, get_income, update_income,
    delete_income, get_income_by_month,
    load_budgets, add_budget, get_budget, update_budget, delete_budget,
    load_recurring_expenses, add_recurring_expense, get_recurring_expense,
    update_recurring_expense, delete_recurring_expense, process_recurring_expenses
)
from .llm import parse_expense_with_llm, get_categories, parse_income_with_llm, get_income_sources

# Load environment variables from .env file
load_dotenv()

app = FastAPI(title="Expense Tracker API", version="1.0.0")


def verify_auth(x_auth_password: Optional[str] = Header(None)):
    """Verify the auth password for protected routes."""
    auth_password = os.getenv("AUTH_PASSWORD")
    
    # If no password is set, allow access (for local development)
    if not auth_password:
        return True
    
    if x_auth_password != auth_password:
        raise HTTPException(
            status_code=401, 
            detail="Unauthorized. Invalid or missing password."
        )
    return True


def get_payment_method(item: dict) -> str:
    method = str(item.get("payment_method") or "debit").lower()
    return method if method in {"debit", "credit"} else "debit"


def format_payment_method(method: str) -> str:
    return "Credit" if method == "credit" else "Debit"

# CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
async def root():
    return {"message": "Expense Tracker API", "version": "1.0.0"}


@app.get("/auth/check")
async def check_auth():
    """Check if authentication is required."""
    auth_password = os.getenv("AUTH_PASSWORD")
    return {"auth_required": bool(auth_password)}


@app.post("/auth/verify")
async def verify_password(authorized: bool = Depends(verify_auth)):
    """Verify the provided password."""
    return {"success": True}


@app.get("/categories")
async def list_categories():
    """Get all available expense categories."""
    return {"categories": get_categories()}


@app.post("/expenses")
async def create_expense(expense: ExpenseCreate, authorized: bool = Depends(verify_auth)):
    """Create a new expense from raw input, using LLM to parse and categorize."""
    api_key = os.getenv("OPENAI_API_KEY")
    
    if not api_key:
        raise HTTPException(
            status_code=400, 
            detail="OPENAI_API_KEY environment variable is not set. Please set it in your .env file."
        )
    
    try:
        # Parse the raw input using LLM
        parsed = parse_expense_with_llm(expense.raw_input, api_key)
        
        # Save to storage
        expense_data = {
            "raw_input": expense.raw_input,
            **parsed
        }
        saved = add_expense(expense_data)
        
        return {"success": True, "expense": saved}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/expenses")
async def list_expenses(year: int = None, month: int = None):
    """Get all expenses, optionally filtered by month."""
    if year and month:
        expenses = get_expenses_by_month(year, month)
    else:
        expenses = load_expenses()
    
    # Sort by date descending
    expenses.sort(key=lambda x: x.get("date", ""), reverse=True)
    
    return {"expenses": expenses}


@app.get("/expenses/{expense_id}")
async def get_single_expense(expense_id: str):
    """Get a single expense by ID."""
    expense = get_expense(expense_id)
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"expense": expense}


@app.put("/expenses/{expense_id}")
async def update_single_expense(expense_id: str, update: ExpenseUpdate, authorized: bool = Depends(verify_auth)):
    """Update an expense."""
    updated = update_expense(expense_id, update.model_dump())
    if not updated:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"success": True, "expense": updated}


@app.delete("/expenses/{expense_id}")
async def delete_single_expense(expense_id: str, authorized: bool = Depends(verify_auth)):
    """Delete an expense."""
    if delete_expense(expense_id):
        return {"success": True}
    raise HTTPException(status_code=404, detail="Expense not found")


# ============ Income Endpoints ============

@app.get("/income/sources")
async def list_income_sources():
    """Get all available income sources."""
    return {"sources": get_income_sources()}


@app.post("/income")
async def create_income(income: IncomeCreate, authorized: bool = Depends(verify_auth)):
    """Create a new income entry from raw input, using LLM to parse."""
    api_key = os.getenv("OPENAI_API_KEY")
    
    if not api_key:
        raise HTTPException(
            status_code=400, 
            detail="OPENAI_API_KEY environment variable is not set."
        )
    
    try:
        parsed = parse_income_with_llm(income.raw_input, api_key)
        
        income_data = {
            "raw_input": income.raw_input,
            **parsed
        }
        saved = add_income(income_data)
        
        return {"success": True, "income": saved}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/income")
async def list_income(year: int = None, month: int = None):
    """Get all income entries, optionally filtered by month."""
    if year and month:
        incomes = get_income_by_month(year, month)
    else:
        incomes = load_incomes()
    
    incomes.sort(key=lambda x: x.get("date", ""), reverse=True)
    
    return {"income": incomes}


@app.get("/income/{income_id}")
async def get_single_income(income_id: str):
    """Get a single income entry by ID."""
    income = get_income(income_id)
    if not income:
        raise HTTPException(status_code=404, detail="Income not found")
    return {"income": income}


@app.put("/income/{income_id}")
async def update_single_income(income_id: str, update: IncomeUpdate, authorized: bool = Depends(verify_auth)):
    """Update an income entry."""
    updated = update_income(income_id, update.model_dump())
    if not updated:
        raise HTTPException(status_code=404, detail="Income not found")
    return {"success": True, "income": updated}


@app.delete("/income/{income_id}")
async def delete_single_income(income_id: str, authorized: bool = Depends(verify_auth)):
    """Delete an income entry."""
    if delete_income(income_id):
        return {"success": True}
    raise HTTPException(status_code=404, detail="Income not found")


@app.get("/analytics")
async def get_analytics(year: int = None, month: int = None):
    """Get spending analytics for a specific month or all time."""
    if year and month:
        expenses = get_expenses_by_month(year, month)
        incomes = get_income_by_month(year, month)
        month_label = f"{year}-{month:02d}"
    else:
        expenses = load_expenses()
        incomes = load_incomes()
        month_label = "all"
    
    # Calculate income totals
    total_income = sum(i.get("amount", 0) for i in incomes)
    by_source = defaultdict(float)
    for income in incomes:
        source = income.get("source", "Other")
        by_source[source] += income.get("amount", 0)
    
    if not expenses and not incomes:
        return {
            "month": month_label,
            "total_spent": 0,
            "total_income": 0,
            "debit_spent": 0,
            "credit_spent": 0,
            "credit_card_balance": 0,
            "current_bank_balance": 0,
            "net_balance": 0,
            "by_category": {},
            "by_payment_method": {"debit": 0, "credit": 0},
            "by_source": {},
            "daily_spending": [],
            "top_categories": [],
            "expense_count": 0,
            "income_count": 0
        }
    
    # Calculate expense totals by category
    by_category = defaultdict(float)
    by_payment_method = defaultdict(float)
    daily_spending = defaultdict(float)
    
    for expense in expenses:
        amount = expense.get("amount", 0)
        category = expense.get("category", "Other")
        payment_method = get_payment_method(expense)
        date = expense.get("date", "")
        
        by_category[category] += amount
        by_payment_method[payment_method] += amount
        if date:
            daily_spending[date] += amount
    
    # Sort categories by amount
    sorted_categories = sorted(by_category.items(), key=lambda x: x[1], reverse=True)
    top_categories = [{"category": cat, "amount": round(amt, 2)} for cat, amt in sorted_categories[:5]]
    
    # Format daily spending for chart
    daily_list = [{"date": date, "amount": round(amt, 2)} for date, amt in sorted(daily_spending.items())]
    
    total_spent = sum(by_category.values())
    debit_spent = by_payment_method.get("debit", 0)
    credit_spent = by_payment_method.get("credit", 0)
    credit_card_balance = credit_spent
    current_bank_balance = total_income - debit_spent
    net_balance = current_bank_balance - credit_card_balance
    
    return {
        "month": month_label,
        "total_spent": round(total_spent, 2),
        "total_income": round(total_income, 2),
        "debit_spent": round(debit_spent, 2),
        "credit_spent": round(credit_spent, 2),
        "credit_card_balance": round(credit_card_balance, 2),
        "current_bank_balance": round(current_bank_balance, 2),
        "net_balance": round(net_balance, 2),
        "by_category": {k: round(v, 2) for k, v in by_category.items()},
        "by_payment_method": {
            "debit": round(debit_spent, 2),
            "credit": round(credit_spent, 2)
        },
        "by_source": {k: round(v, 2) for k, v in by_source.items()},
        "daily_spending": daily_list,
        "top_categories": top_categories,
        "expense_count": len(expenses),
        "income_count": len(incomes)
    }


@app.get("/analytics/months")
async def get_available_months():
    """Get all months that have expense data."""
    months = get_all_months()
    return {"months": months}


@app.get("/analytics/trends")
async def get_trends():
    """Get spending trends across all months."""
    months = get_all_months()
    trends = []
    
    for month_str in months:
        year, month = map(int, month_str.split("-"))
        expenses = get_expenses_by_month(year, month)
        
        by_category = defaultdict(float)
        by_payment_method = defaultdict(float)
        total = 0
        
        for expense in expenses:
            amount = expense.get("amount", 0)
            category = expense.get("category", "Other")
            payment_method = get_payment_method(expense)
            by_category[category] += amount
            by_payment_method[payment_method] += amount
            total += amount
        
        trends.append({
            "month": month_str,
            "total": round(total, 2),
            "debit_spent": round(by_payment_method.get("debit", 0), 2),
            "credit_spent": round(by_payment_method.get("credit", 0), 2),
            "by_category": {k: round(v, 2) for k, v in by_category.items()},
            "expense_count": len(expenses)
        })
    
    return {"trends": trends}


# ============ Budget Endpoints ============

@app.get("/budgets")
async def list_budgets():
    """Get all budgets."""
    budgets = load_budgets()
    return {"budgets": budgets}


@app.post("/budgets")
async def create_budget(budget: BudgetCreate, authorized: bool = Depends(verify_auth)):
    """Create a new budget."""
    budget_data = {
        "category": budget.category,
        "monthly_limit": budget.monthly_limit
    }
    saved = add_budget(budget_data)
    return {"success": True, "budget": saved}


@app.get("/budgets/{budget_id}")
async def get_single_budget(budget_id: str):
    """Get a single budget by ID."""
    budget = get_budget(budget_id)
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")
    return {"budget": budget}


@app.put("/budgets/{budget_id}")
async def update_single_budget(budget_id: str, update: BudgetUpdate, authorized: bool = Depends(verify_auth)):
    """Update a budget."""
    updated = update_budget(budget_id, update.model_dump())
    if not updated:
        raise HTTPException(status_code=404, detail="Budget not found")
    return {"success": True, "budget": updated}


@app.delete("/budgets/{budget_id}")
async def delete_single_budget(budget_id: str, authorized: bool = Depends(verify_auth)):
    """Delete a budget."""
    if delete_budget(budget_id):
        return {"success": True}
    raise HTTPException(status_code=404, detail="Budget not found")


@app.get("/budgets/status/{year}/{month}")
async def get_budget_status(year: int, month: int):
    """Get budget status with spending progress for a specific month."""
    budgets = load_budgets()
    expenses = get_expenses_by_month(year, month)
    
    # Calculate spending by category
    spending_by_category = defaultdict(float)
    for expense in expenses:
        category = expense.get("category", "Other")
        spending_by_category[category] += expense.get("amount", 0)
    
    # Build status for each budget
    status = []
    for budget in budgets:
        category = budget["category"]
        limit = budget["monthly_limit"]
        spent = spending_by_category.get(category, 0)
        percentage = (spent / limit * 100) if limit > 0 else 0
        
        status.append({
            "id": budget["id"],
            "category": category,
            "monthly_limit": limit,
            "spent": round(spent, 2),
            "remaining": round(limit - spent, 2),
            "percentage": round(percentage, 1),
            "status": "exceeded" if percentage >= 100 else "warning" if percentage >= 80 else "ok"
        })
    
    return {"budget_status": status, "month": f"{year}-{month:02d}"}


# ============ Recurring Expense Endpoints ============

@app.get("/recurring")
async def list_recurring_expenses():
    """Get all recurring expenses."""
    recurring = load_recurring_expenses()
    return {"recurring": recurring}


@app.post("/recurring")
async def create_recurring_expense(recurring: RecurringExpenseCreate, authorized: bool = Depends(verify_auth)):
    """Create a new recurring expense."""
    recurring_data = {
        "description": recurring.description,
        "amount": recurring.amount,
        "category": recurring.category,
        "payment_method": recurring.payment_method,
        "day_of_month": recurring.day_of_month
    }
    saved = add_recurring_expense(recurring_data)
    return {"success": True, "recurring": saved}


@app.get("/recurring/{recurring_id}")
async def get_single_recurring_expense(recurring_id: str):
    """Get a single recurring expense by ID."""
    recurring = get_recurring_expense(recurring_id)
    if not recurring:
        raise HTTPException(status_code=404, detail="Recurring expense not found")
    return {"recurring": recurring}


@app.put("/recurring/{recurring_id}")
async def update_single_recurring_expense(recurring_id: str, update: RecurringExpenseUpdate, authorized: bool = Depends(verify_auth)):
    """Update a recurring expense."""
    updated = update_recurring_expense(recurring_id, update.model_dump())
    if not updated:
        raise HTTPException(status_code=404, detail="Recurring expense not found")
    return {"success": True, "recurring": updated}


@app.delete("/recurring/{recurring_id}")
async def delete_single_recurring_expense(recurring_id: str, authorized: bool = Depends(verify_auth)):
    """Delete a recurring expense."""
    if delete_recurring_expense(recurring_id):
        return {"success": True}
    raise HTTPException(status_code=404, detail="Recurring expense not found")


@app.post("/recurring/process")
async def trigger_recurring_processing(authorized: bool = Depends(verify_auth)):
    """Manually trigger processing of recurring expenses."""
    added = process_recurring_expenses()
    return {"success": True, "added_count": len(added), "added_expenses": added}


# ============ Export Endpoint ============

@app.get("/export/excel")
async def export_excel():
    """Export all data (expenses, income, budgets, recurring) as an Excel file."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill_green = PatternFill(start_color="1A7A4A", end_color="1A7A4A", fill_type="solid")
    header_fill_blue = PatternFill(start_color="1A4A7A", end_color="1A4A7A", fill_type="solid")
    header_fill_purple = PatternFill(start_color="5A1A7A", end_color="5A1A7A", fill_type="solid")
    header_fill_orange = PatternFill(start_color="7A4A1A", end_color="7A4A1A", fill_type="solid")

    def style_header_row(ws, fill):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    # --- Expenses Sheet ---
    expenses = load_expenses()
    expenses.sort(key=lambda x: x.get("date", ""), reverse=True)
    ws_exp = wb.active
    ws_exp.title = "Expenses"
    ws_exp.append(["Date", "Description", "Amount ($)", "Category", "Payment Method"])
    style_header_row(ws_exp, header_fill_green)
    for e in expenses:
        ws_exp.append([
            e.get("date", ""),
            e.get("description", ""),
            e.get("amount", 0),
            e.get("category", ""),
            format_payment_method(get_payment_method(e))
        ])
    auto_width(ws_exp)

    # --- Income Sheet ---
    incomes = load_incomes()
    incomes.sort(key=lambda x: x.get("date", ""), reverse=True)
    ws_inc = wb.create_sheet("Income")
    ws_inc.append(["Date", "Description", "Amount ($)", "Source"])
    style_header_row(ws_inc, header_fill_blue)
    for i in incomes:
        ws_inc.append([
            i.get("date", ""),
            i.get("description", ""),
            i.get("amount", 0),
            i.get("source", "")
        ])
    auto_width(ws_inc)

    # --- Budgets Sheet ---
    budgets = load_budgets()
    ws_bud = wb.create_sheet("Budgets")
    ws_bud.append(["Category", "Monthly Limit ($)"])
    style_header_row(ws_bud, header_fill_purple)
    for b in budgets:
        ws_bud.append([b.get("category", ""), b.get("monthly_limit", 0)])
    auto_width(ws_bud)

    # --- Recurring Sheet ---
    recurring = load_recurring_expenses()
    ws_rec = wb.create_sheet("Recurring")
    ws_rec.append(["Description", "Amount ($)", "Category", "Payment Method", "Day of Month", "Active"])
    style_header_row(ws_rec, header_fill_orange)
    for r in recurring:
        ws_rec.append([
            r.get("description", ""),
            r.get("amount", 0),
            r.get("category", ""),
            format_payment_method(get_payment_method(r)),
            r.get("day_of_month", ""),
            "Yes" if r.get("is_active", True) else "No"
        ])
    auto_width(ws_rec)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"expense_tracker_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
